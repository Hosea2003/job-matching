from typing import Any
from django.core.management import BaseCommand
import openpyxl
from django.conf import settings
import os
import csv

from matching.models import (
    HardSkill,
    JobHardSkill,
    JobOffer,
    JobSoftSkill,
    SoftSkill,
    User,
    UserHardSkill,
    UserSoftSkill,
)


class Command(BaseCommand):
    def handle(self, *args: Any, **options: Any) -> str | None:
        # self.load_user()
        # self.load_job_offer()
        # self.load_hard_skills()
        # self.load_soft_skills()
        # self.load_job_skills()
        self.load_user_skills()
        pass

    def load_user(self):
        dataframe = openpyxl.load_workbook(
            os.path.join(settings.BASE_DIR, "files/USERS.xlsx")
        )

        user_frame = dataframe.active
        users = []

        for row in user_frame.iter_rows(2, user_frame.max_row):
            id = int(row[0].value)
            email = row[1].value
            first_name = row[2].value
            last_name = row[3].value
            experience = int(row[4].value)

            users.append(
                User(
                    id=id,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    experience=experience,
                )
            )

        User.objects.bulk_create(users)

    def load_job_offer(self):
        dataframe = openpyxl.load_workbook(
            os.path.join(settings.BASE_DIR, "files/JOB OFFER.xlsx")
        )

        job_offer_frame = dataframe.active
        job_offer = []

        for row in job_offer_frame.iter_rows(2, job_offer_frame.max_row):
            id = int(row[0].value)
            name = row[1].value
            experience = int(row[2].value)

            job_offer.append(
                JobOffer(
                    id=id,
                    name=name,
                    experience=experience,
                )
            )

        JobOffer.objects.bulk_create(job_offer)

    def load_hard_skills(self):
        dataframe = openpyxl.load_workbook(
            os.path.join(settings.BASE_DIR, "files/HARD SKILLS.xlsx")
        )

        hard_skills_frame = dataframe.active
        hard_skills = []

        for row in hard_skills_frame.iter_rows(2, hard_skills_frame.max_row):
            id = int(row[0].value)
            name = row[1].value

            hard_skills.append(
                HardSkill(
                    id=id,
                    name=name,
                )
            )

        HardSkill.objects.bulk_create(hard_skills)

    def load_soft_skills(self):
        dataframe = openpyxl.load_workbook(
            os.path.join(settings.BASE_DIR, "files/SOFT SKILLS.xlsx")
        )

        soft_skills_frame = dataframe.active
        soft_skills = []

        for row in soft_skills_frame.iter_rows(2, soft_skills_frame.max_row):
            id = int(row[0].value)
            name = row[1].value

            soft_skills.append(
                SoftSkill(
                    id=id,
                    name=name,
                )
            )

        SoftSkill.objects.bulk_create(soft_skills)

    def load_job_skills(self):
        with open(
            os.path.join(settings.BASE_DIR, "files/job_hard_skills_filtered.csv"),
            mode="r",
        ) as file:
            job_skills = csv.DictReader(file)
            job_hard_skills = []
            for line in job_skills:
                row = next(iter(line.values())).split(",")
                job_id = int(row[0])
                hard_skill_id = int(row[1])
                score = int(row[2])

                job_hard_skills.append(
                    JobHardSkill(job_id=job_id, hardskill_id=hard_skill_id, level=score)
                )

            JobHardSkill.objects.bulk_create(job_hard_skills)

        with open(
            os.path.join(settings.BASE_DIR, "files/job_soft_skills_filtered.csv"),
            mode="r",
        ) as file:
            job_skills = csv.DictReader(file)
            job_soft_skills = []
            for line in job_skills:
                job_id = int(line["Job ID"])
                soft_skill_id = int(line["Soft Skill ID"])
                score = int(line["Score"])

                job_soft_skills.append(
                    JobSoftSkill(job_id=job_id, softskill_id=soft_skill_id, level=score)
                )

            JobSoftSkill.objects.bulk_create(job_soft_skills)

    def load_user_skills(self):
        with open(
            os.path.join(settings.BASE_DIR, "files/user_hard_skills.csv"), "r"
        ) as file:
            user_skills_data = csv.DictReader(file)
            user_skills = []
            for line in user_skills_data:
                user_id = int(line["User ID"])
                hard_skill_id = int(line["Hard Skill ID"])
                level = int(line["Score"])

                user_skills.append(
                    UserHardSkill(
                        user_id=user_id, hardskill_id=hard_skill_id, level=level
                    )
                )
            UserHardSkill.objects.bulk_create(user_skills)

        with open(
            os.path.join(settings.BASE_DIR, "files/user_soft_skills.csv"), "r"
        ) as file:
            user_skills_data = csv.DictReader(file)
            user_skills = []
            for line in user_skills_data:
                user_id = int(line["User ID"])
                soft_skill_id = int(line["Soft Skill ID"])
                level = int(line["Score"])

                user_skills.append(
                    UserSoftSkill(
                        user_id=user_id, softskill_id=soft_skill_id, level=level
                    )
                )
            UserSoftSkill.objects.bulk_create(user_skills)
